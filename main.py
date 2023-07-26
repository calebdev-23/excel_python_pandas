import os.path

import openpyxl
import pandas as pd

# Version 1
'''wb = openpyxl.load_workbook("octobre.xlsx", data_only=True)
sheet = wb.active
row = sheet.max_row
cel = sheet.max_column
column = 0
data = {}
for i in range(2, row):
    for j in range(2, cel+1):
        if not sheet.cell(i, j).value:
            break
        else:
            if data.get(sheet.cell(i, 1).value):
                data[sheet.cell(i, 1).value].append(sheet.cell(i, j).value)
            else:
                data[sheet.cell(i, 1).value] = [sheet.cell(i, j).value]

print(data) '''

# Version 2
'''wb = openpyxl.load_workbook('novembre.xlsx')
sheet = wb.active
columns = sheet.max_column
rows = sheet.max_row
listes = []
print("nombre de colonne :", columns, "nombre ligne :", rows)
print()
for ligne in range(1, columns):
    # print(sheet.cell(1, ligne).value)
    # print(sheet.cell(row, ligne).value)
    for row in range(2, rows):
        if not sheet.cell(row, ligne).value:
            break
        dic = {sheet.cell(1, ligne).value: sheet.cell(row, ligne).value}
        listes.append(dic)
listes.reverse()
for liste in listes:
    print(liste)
if os.path.exists("donne.xlsx"):
    os.remove("donne.xlsx")

df = pd.DataFrame.from_records(listes)
print(df)
#df.to_excel("donne.xlsx", index=False)'''

# Version 3

'''wb = openpyxl.load_workbook('novembre.xlsx', data_only=True)
sheet = wb.active
columns = sheet.max_column
rows = sheet.max_row
listes = []
print("nombre de colonne :", columns, "nombre ligne :", rows)
print()
for row in range(1, rows):
    # print(sheet.cell(1, ligne).value)
    # print(sheet.cell(row, col).value)
    for col in range(1, columns + 1):
        if not sheet.cell(row, col).value:
            break
        dic = {
            sheet.cell(1, col).value: sheet.cell(row+1, col).value
        }
        listes.append(dic)

df = pd.DataFrame.from_records(listes)
print(df) '''

# Version 4

wb = openpyxl.load_workbook('novembre.xlsx', data_only=True)
sheet = wb.active
columns = sheet.max_column
rows = sheet.max_row
listes = []
print("nombre de colonne :", columns, "nombre ligne :", rows)
print()
dic = {}
for col in range(1, columns+1):
    for row in range(2, rows):
        if not sheet.cell(row, col).value:
            break
        else:
            if dic.get(sheet.cell(1, col).value):
                dic[sheet.cell(1, col).value].append(sheet.cell(row, col).value)
            else:
                dic[sheet.cell(1, col).value] = [sheet.cell(row, col).value]

df = pd.DataFrame(dic)
if os.path.exists("donne.xlsx"):
    os.remove("donne.xlsx")
df.to_excel('donne.xlsx', index=False)


